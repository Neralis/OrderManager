from datetime import datetime
from io import BytesIO

import openpyxl
import qrcode
from django.utils.dateparse import parse_date
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from orderApp.models import Order, OrderItem
from productApp.models import Stock
from openpyxl.drawing.image import Image as OpenpyxlImage


def export_filtered_orders_to_xlsx(request):
    start_raw = request.GET.get("start")
    end_raw = request.GET.get("end")

    start = parse_date(start_raw) if start_raw else None
    end = parse_date(end_raw) if end_raw else None

    orders = Order.objects.prefetch_related('items__product').select_related('warehouse')
    if start:
        orders = orders.filter(created_at__date__gte=start)
    if end:
        orders = orders.filter(created_at__date__lte=end)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"

    headers = [
        "ID заказа", "Статус", "Склад", "Дата создания",
        "Клиент", "Адрес", "Комментарий", "Причина отмены",
        "Продукт", "Количество", "Цена за единицу", "Сумма по товару", "Итоговая сумма заказа"
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    status_colors = {
        "new": "FFFACD",
        "processing": "ADD8E6",
        "shipped": "87CEEB",
        "completed": "90EE90",
        "cancelled": "FFC0CB",
    }

    for order in orders:
        for item in order.items.all():
            sheet.append([
                order.id,
                order.status,
                order.warehouse.name if order.warehouse else "—",
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                order.client_name,
                order.destination_address,
                order.comment,
                order.cancellation_reason,
                item.product.name,
                item.quantity,
                float(item.price),
                float(item.price) * item.quantity,
                float(order.total_price)
            ])

            # Цветовая заливка по статусу
            row_num = sheet.max_row
            status_cell = sheet.cell(row=row_num, column=2)
            fill_color = status_colors.get(order.status)
            if fill_color:
                status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Автоширина
    for col in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_len + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=orders_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    workbook.save(response)
    return response



def export_stock_to_xlsx(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Остатки товаров"

    headers = ["Склад", "Продукт", "Остаток"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    stocks = Stock.objects.select_related("warehouse", "product").order_by("warehouse__name", "product__name")

    for stock in stocks:
        sheet.append([
            stock.warehouse.name,
            stock.product.name,
            stock.quantity
        ])

    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    workbook.save(response)
    return response


def export_single_order_to_xlsx(request, order_id: int):
    order = Order.objects.prefetch_related('items__product').select_related('warehouse').get(id=order_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ {order.id}"

    # Общие сведения о заказе
    order_info = [
        ("ID заказа", order.id),
        ("Статус", order.status),
        ("Дата создания", order.created_at.strftime("%Y-%m-%d %H:%M")),
        ("Склад", order.warehouse.name if order.warehouse else "—"),
        ("Клиент", order.client_name),
        ("Адрес доставки", order.destination_address),
        ("Комментарий", order.comment or "—"),
        ("Причина отмены", order.cancellation_reason or "—"),
        ("Итоговая сумма", float(order.total_price)),
    ]

    for idx, (label, value) in enumerate(order_info, start=1):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)

    start_row = len(order_info) + 2
    ws.cell(row=start_row, column=1, value="Продукты").font = Font(bold=True)
    headers = ["Продукт", "Количество", "Цена за единицу", "Сумма"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row + 1, column=col, value=header).font = Font(bold=True)

    for idx, item in enumerate(order.items.all(), start=1):
        row = start_row + 1 + idx
        ws.cell(row=row, column=1, value=item.product.name)
        ws.cell(row=row, column=2, value=item.quantity)
        ws.cell(row=row, column=3, value=float(item.price))
        ws.cell(row=row, column=4, value=float(item.price) * item.quantity)

    # QR-код
    qr_text = (
        f"Order ID: {order.id}\nStatus: {order.status}\nClient: {order.client_name}\n"
        f"Total: {order.total_price}\nCreated: {order.created_at.strftime('%Y-%m-%d %H:%M')}"
    )
    qr_img = qrcode.make(qr_text)
    img_io = BytesIO()
    qr_img.save(img_io, format='PNG')
    img_io.seek(0)
    img = OpenpyxlImage(img_io)
    img.width = 150
    img.height = 150
    ws.add_image(img, f"F2")

    # Автоширина
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=order_{order.id}_report.xlsx"
    wb.save(response)
    return response
